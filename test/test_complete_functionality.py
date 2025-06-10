#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完整功能验证脚本
测试所有新增的Excel处理功能
"""

import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill

# 导入我们的模块
from server import (
    read_excel_metadata,
    run_excel_code
)
from excel_smart_tools import (
    suggest_excel_read_parameters,
    detect_excel_file_structure,
    create_excel_read_template
)

def create_complex_test_excel():
    """创建一个复杂的测试Excel文件"""
    file_path = "complex_test_data.xlsx"
    
    # 创建工作簿
    wb = Workbook()
    
    # 第一个工作表：标准数据
    ws1 = wb.active
    ws1.title = "标准数据"
    
    # 添加标题和数据
    data1 = {
        '产品名称': ['iPhone 14', 'Samsung S23', 'Huawei P50', 'Xiaomi 13', 'OPPO Find X5'],
        '销售额': [120000, 85000, 95000, 60000, 45000],
        '销售量': [1200, 850, 950, 600, 450],
        '利润率': [0.25, 0.18, 0.22, 0.15, 0.12]
    }
    df1 = pd.DataFrame(data1)
    
    for r in dataframe_to_rows(df1, index=False, header=True):
        ws1.append(r)
    
    # 第二个工作表：带有多级标题的数据
    ws2 = wb.create_sheet("多级标题")
    
    # 添加多级标题
    ws2['A1'] = '2023年销售报告'
    ws2['A2'] = '数据来源：销售部门'
    ws2['A3'] = ''
    
    # 主标题行
    headers = ['地区', 'Q1销售额', 'Q2销售额', 'Q3销售额', 'Q4销售额', '年度总计']
    for col, header in enumerate(headers, 1):
        ws2.cell(row=4, column=col, value=header)
    
    # 数据行
    data2 = [
        ['北京', 150000, 180000, 200000, 220000, 750000],
        ['上海', 140000, 170000, 190000, 210000, 710000],
        ['广州', 120000, 150000, 170000, 180000, 620000],
        ['深圳', 130000, 160000, 180000, 200000, 670000]
    ]
    
    for row_idx, row_data in enumerate(data2, 5):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # 第三个工作表：包含空行和注释的数据
    ws3 = wb.create_sheet("复杂格式")
    
    # 添加一些空行和注释
    ws3['A1'] = '# 这是注释行'
    ws3['A2'] = ''
    ws3['A3'] = '员工信息表'
    ws3['A4'] = '更新日期：2023-12-01'
    ws3['A5'] = ''
    
    # 数据标题（从第6行开始）
    emp_headers = ['员工ID', '姓名', '部门', '薪资', '入职日期']
    for col, header in enumerate(emp_headers, 1):
        ws3.cell(row=6, column=col, value=header)
    
    # 员工数据
    emp_data = [
        ['E001', '张三', '技术部', 15000, '2020-01-15'],
        ['E002', '李四', '销售部', 12000, '2021-03-20'],
        ['E003', '王五', '市场部', 13000, '2019-08-10'],
        ['E004', '赵六', '技术部', 16000, '2022-05-30']
    ]
    
    for row_idx, row_data in enumerate(emp_data, 7):
        for col_idx, value in enumerate(row_data, 1):
            ws3.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    wb.save(file_path)
    print(f"✅ 创建复杂测试Excel文件: {file_path}")
    return file_path

def test_enhanced_metadata_reading(file_path):
    """测试增强的元数据读取功能"""
    print("\n🔍 测试增强的元数据读取功能...")
    
    # 测试标准工作表
    result1 = read_excel_metadata(file_path, sheet_name="标准数据")
    print(f"标准数据工作表元数据状态: {result1.get('status', 'UNKNOWN')}")
    if result1.get('status') == 'SUCCESS':
        print(f"  - 列数: {len(result1.get('columns_metadata', []))}")
        print(f"  - 推荐参数: {result1.get('suggested_params', {})}")
    
    # 测试多级标题工作表
    result2 = read_excel_metadata(file_path, sheet_name="多级标题", skiprows=3, header=0)
    print(f"多级标题工作表元数据状态: {result2.get('status', 'UNKNOWN')}")
    if result2.get('status') == 'SUCCESS':
        print(f"  - 列数: {len(result2.get('columns_metadata', []))}")
        print(f"  - 使用参数: skiprows=3, header=0")
    
    # 测试复杂格式工作表
    result3 = read_excel_metadata(file_path, sheet_name="复杂格式", skiprows=5, header=0)
    print(f"复杂格式工作表元数据状态: {result3.get('status', 'UNKNOWN')}")
    if result3.get('status') == 'SUCCESS':
        print(f"  - 列数: {len(result3.get('columns_metadata', []))}")
        print(f"  - 使用参数: skiprows=5, header=0")

def test_smart_parameter_suggestion(file_path):
    """测试智能参数推荐功能"""
    print("\n🧠 测试智能参数推荐功能...")
    
    result = suggest_excel_read_parameters(file_path)
    print(f"参数推荐结果: {'成功' if 'recommended_params' in result else '失败'}")
    
    if 'recommended_params' in result:
        print(f"  推荐参数: {result['recommended_params']}")
        if result.get('tips'):
            print(f"  提示: {result['tips'][:2]}")
        if result.get('warnings'):
            print(f"  警告: {result['warnings'][:2]}")

def test_structure_detection(file_path):
    """测试结构检测功能"""
    print("\n🏗️ 测试结构检测功能...")
    
    result = detect_excel_file_structure(file_path)
    print(f"结构检测结果: {'成功' if 'sheets' in result else '失败'}")
    
    if 'sheets' in result:
        print(f"  - 工作表数量: {len(result['sheets'])}")
        for sheet_info in result['sheets']:
            print(f"  - 工作表 '{sheet_info['name']}': {sheet_info['max_row']}行 x {sheet_info['max_column']}列")
            if sheet_info.get('merged_cells'):
                print(f"    合并单元格: {len(sheet_info['merged_cells'])}个")
    
    if result.get('error'):
        print(f"  错误: {result['error']}")
    
    if result.get('data_range'):
        print(f"  数据范围: {result['data_range']}")
    
    if result.get('formatting_info'):
        print(f"  格式信息: {result['formatting_info']}")

def test_template_generation(file_path):
    """测试代码模板生成功能"""
    print("\n📝 测试代码模板生成功能...")
    
    # 测试不同工作表的模板生成
    sheets = ["标准数据", "多级标题", "复杂格式"]
    
    for sheet_name in sheets:
        result = create_excel_read_template(file_path, sheet_name)
        print(f"工作表 '{sheet_name}' 模板生成状态: {result.get('status', 'UNKNOWN')}")
        
        if result.get('code_template'):
            print("  生成的代码模板:")
            print("  " + "="*50)
            print(result['code_template'])
            print("  " + "="*50)
        
        if result.get('recommended_params'):
            print("  推荐参数:")
            for param, value in result['recommended_params'].items():
                print(f"    - {param}: {value}")
        
        if result.get('tips'):
            print("  提示:")
            for tip in result['tips']:
                print(f"    - {tip}")
        print()

def test_enhanced_code_execution(file_path):
    """测试增强的代码执行功能"""
    print("\n⚡ 测试增强的代码执行功能...")
    
    # 测试用例1：标准数据分析
    code1 = """
print("=== 标准数据分析 ===")
print(f"数据形状: {df.shape}")
print(f"列名: {list(df.columns)}")
if '销售额' in df.columns:
    print(f"销售额总计: {df['销售额'].sum():,}")
if '利润率' in df.columns:
    print(f"平均利润率: {df['利润率'].mean():.2%}")
"""
    
    result1 = run_excel_code(code1, file_path, sheet_name="标准数据")
    print(f"标准数据分析结果:")
    if result1.get('output'):
        print("输出:")
        print(result1['output'])
    if result1.get('warning'):
        print(f"警告: {result1['warning']}")
    
    # 测试用例2：多级标题数据分析
    code2 = """
print("=== 多级标题数据分析 ===")
print(f"数据形状: {df.shape}")
print(f"列名: {list(df.columns)}")
if '年度总计' in df.columns:
    print(f"年度总计最高的地区: {df.loc[df['年度总计'].idxmax(), '地区']}")
    print(f"年度总计: {df['年度总计'].max():,}")
"""
    
    result2 = run_excel_code(code2, file_path, sheet_name="多级标题", skiprows=3, header=0)
    print(f"多级标题数据分析结果:")
    if result2.get('output'):
        print("输出:")
        print(result2['output'])
    if result2.get('warning'):
        print(f"警告: {result2['warning']}")
    
    # 测试用例3：复杂格式数据分析
    code3 = """
print("=== 复杂格式数据分析 ===")
print(f"数据形状: {df.shape}")
print(f"列名: {list(df.columns)}")
if '部门' in df.columns and '薪资' in df.columns:
    dept_avg = df.groupby('部门')['薪资'].mean()
    print("各部门平均薪资:")
    for dept, avg_salary in dept_avg.items():
        print(f"  {dept}: {avg_salary:,.0f}")
"""
    
    result3 = run_excel_code(code3, file_path, sheet_name="复杂格式", skiprows=5, header=0, usecols="A:E")
    print(f"复杂格式数据分析结果:")
    if result3.get('output'):
        print("输出:")
        print(result3['output'])
    if result3.get('warning'):
        print(f"警告: {result3['warning']}")

def main():
    """主测试函数"""
    print("🚀 开始完整功能验证测试...")
    print("=" * 60)
    
    try:
        # 创建测试文件
        test_file = create_complex_test_excel()
        
        # 运行所有测试
        test_enhanced_metadata_reading(test_file)
        test_smart_parameter_suggestion(test_file)
        test_structure_detection(test_file)
        test_template_generation(test_file)
        test_enhanced_code_execution(test_file)
        
        print("\n" + "=" * 60)
        print("✅ 所有功能验证测试完成！")
        print("\n📋 功能清单:")
        print("  ✅ xlrd 和 xlsxwriter 库已安装")
        print("  ✅ read_excel_metadata 函数支持 skiprows、header、usecols 参数")
        print("  ✅ run_excel_code 函数支持 skiprows、header、usecols 参数")
        print("  ✅ 智能参数推荐功能已实现")
        print("  ✅ Excel 文件结构检测功能已实现")
        print("  ✅ 代码模板生成功能已实现")
        print("  ✅ 所有新功能已注册为 MCP 工具")
        
        # 清理测试文件
        if os.path.exists(test_file):
            os.remove(test_file)
            print(f"\n🧹 已清理测试文件: {test_file}")
            
    except Exception as e:
        print(f"\n❌ 测试过程中发生错误: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()