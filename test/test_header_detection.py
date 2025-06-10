#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试标题行检测逻辑
"""

import pandas as pd
import os
from excel_helper import _suggest_excel_read_parameters

def create_simple_test_file():
    """创建简单的测试文件"""
    test_data = {
        '姓名': ['张三', '李四', '王五', '赵六', '钱七'],
        '年龄': [25, 30, 35, 28, 32],
        '城市': ['北京', '上海', '广州', '深圳', '杭州'],
        '薪资': [8000, 12000, 15000, 9000, 11000],
        '部门': ['技术', '销售', '市场', '技术', '财务']
    }
    
    df = pd.DataFrame(test_data)
    excel_file = 'simple_test.xlsx'
    df.to_excel(excel_file, index=False)
    return excel_file

def create_complex_test_file():
    """创建复杂的测试文件（带空行和多个潜在标题行）"""
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # 添加一些空行
    ws.cell(row=1, column=1, value=None)
    ws.cell(row=2, column=1, value=None)
    
    # 添加一个假的标题行
    ws.cell(row=3, column=1, value="员工信息表")
    ws.cell(row=3, column=2, value="2024年")
    
    # 添加另一个假的标题行
    ws.cell(row=4, column=1, value="部门")
    ws.cell(row=4, column=2, value="个人信息")
    ws.cell(row=4, column=3, value="薪资信息")
    
    # 真正的列头
    ws.cell(row=5, column=1, value="姓名")
    ws.cell(row=5, column=2, value="年龄")
    ws.cell(row=5, column=3, value="城市")
    ws.cell(row=5, column=4, value="薪资")
    ws.cell(row=5, column=5, value="部门")
    
    # 数据行
    data = [
        ['张三', 25, '北京', 8000, '技术'],
        ['李四', 30, '上海', 12000, '销售'],
        ['王五', 35, '广州', 15000, '市场'],
        ['赵六', 28, '深圳', 9000, '技术'],
        ['钱七', 32, '杭州', 11000, '财务']
    ]
    
    for i, row_data in enumerate(data, start=6):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)
    
    excel_file = 'complex_test.xlsx'
    wb.save(excel_file)
    return excel_file

def test_header_detection(excel_file, file_description):
    """测试标题行检测"""
    print(f"\n🔍 测试文件: {file_description}")
    print("=" * 50)
    
    # 调用参数建议函数
    suggestions = _suggest_excel_read_parameters(excel_file)
    
    print(f"推荐参数: {suggestions['recommended_params']}")
    print(f"分析结果: {suggestions['analysis']}")
    print(f"警告: {suggestions['warnings']}")
    print(f"提示: {suggestions['tips']}")
    
    # 使用推荐参数读取文件
    try:
        df = pd.read_excel(excel_file, **suggestions['recommended_params'])
        print(f"\n读取成功:")
        print(f"  形状: {df.shape}")
        print(f"  列名: {list(df.columns)}")
        print(f"  前几行数据:")
        print(df.head())
        
        # 检查是否包含预期的列
        expected_columns = ['姓名', '年龄', '城市', '薪资', '部门']
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            print(f"  ❌ 缺少预期列: {missing_columns}")
            return False
        else:
            print(f"  ✅ 包含所有预期列")
            return True
            
    except Exception as e:
        print(f"  ❌ 读取失败: {e}")
        return False

def test_manual_parameters(excel_file, file_description):
    """测试手动设置正确参数"""
    print(f"\n🔧 手动参数测试: {file_description}")
    print("=" * 50)
    
    # 对于复杂文件，手动设置正确的参数
    if 'complex' in file_description:
        manual_params = {'header': 4}  # 第5行（0-indexed为4）是真正的列头
    else:
        manual_params = {'header': 0}  # 第1行是列头
    
    try:
        df = pd.read_excel(excel_file, **manual_params)
        print(f"手动参数: {manual_params}")
        print(f"读取成功:")
        print(f"  形状: {df.shape}")
        print(f"  列名: {list(df.columns)}")
        print(f"  前几行数据:")
        print(df.head())
        
        # 检查是否包含预期的列
        expected_columns = ['姓名', '年龄', '城市', '薪资', '部门']
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            print(f"  ❌ 缺少预期列: {missing_columns}")
            return False
        else:
            print(f"  ✅ 包含所有预期列")
            return True
            
    except Exception as e:
        print(f"  ❌ 读取失败: {e}")
        return False

def main():
    """主测试函数"""
    print("🧪 标题行检测测试")
    print("=" * 60)
    
    # 测试简单文件
    simple_file = create_simple_test_file()
    simple_success = test_header_detection(simple_file, "简单文件（标准格式）")
    simple_manual_success = test_manual_parameters(simple_file, "简单文件")
    
    # 测试复杂文件
    complex_file = create_complex_test_file()
    complex_success = test_header_detection(complex_file, "复杂文件（带空行和假标题）")
    complex_manual_success = test_manual_parameters(complex_file, "复杂文件")
    
    # 总结结果
    print("\n" + "=" * 60)
    print("📊 测试总结:")
    print(f"简单文件自动检测: {'✅ 通过' if simple_success else '❌ 失败'}")
    print(f"简单文件手动参数: {'✅ 通过' if simple_manual_success else '❌ 失败'}")
    print(f"复杂文件自动检测: {'✅ 通过' if complex_success else '❌ 失败'}")
    print(f"复杂文件手动参数: {'✅ 通过' if complex_manual_success else '❌ 失败'}")
    
    # 清理文件
    for file in [simple_file, complex_file]:
        if os.path.exists(file):
            os.remove(file)
    
    print(f"\n🧹 已清理测试文件")
    
    # 如果复杂文件的自动检测失败但手动参数成功，说明检测逻辑有问题
    if not complex_success and complex_manual_success:
        print("\n⚠️ 发现问题：复杂文件的自动检测失败，但手动设置正确参数可以成功读取")
        print("   这表明标题行检测逻辑需要改进")
        return False
    
    return simple_success and complex_success

if __name__ == "__main__":
    success = main()
    if not success:
        exit(1)